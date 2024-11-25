import pandas as pd
import re
import os
from openpyxl import load_workbook  # Ensure this is imported for working with existing workbooks

# Constants
MERGE_DROP_LIST = ['BOUNCED', 'ERROR', '0', 'RESPONDED', 'NO_RECIPIENT', 'UNSUBSCRIBED', 'UNINTERESTED']
EXCEL_FILE_PATH = '2023 Email list.xlsx'  # Path to your existing Excel file
NEW_EXCEL_FILE_PATH = '2024 Summit List 1.xlsx'  # Specify the new workbook's name
DO_NOT_EMAIL_FILE_PATH = 'DO NOT EMAIL.xlsx'  # Replace 'Bad Emails.xlsx' with 'DO NOT EMAIL.xlsx'

def is_valid_email(email):
    # Simple regex for validating an email address format
    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    return re.fullmatch(regex, email) is not None

def load_do_not_email_list(path):
    if os.path.exists(path):
        df = pd.read_excel(path)
        return set(df['EMAIL'].str.lower())
    else:
        return set()

def update_do_not_email_list(emails, path):
    if emails.empty:  # If there are no emails to update, return early
        return

    if os.path.exists(path):
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book = book
        if 'DO NOT EMAIL' in book.sheetnames:  # Check if the sheet already exists
            start_row = book['DO NOT EMAIL'].max_row + 1
            header = False  # Do not write header if sheet exists
        else:
            start_row = 0
            header = True  # Write header if sheet does not exist
    else:
        writer = pd.ExcelWriter(path, engine='openpyxl')
        start_row = 0
        header = True  # Write header if creating a new workbook

    new_df = pd.DataFrame(list(emails), columns=['EMAIL'])
    new_df.to_excel(writer, sheet_name='DO NOT EMAIL', index=False, header=header, startrow=start_row)
    writer.save()
    writer.close()

def process_workbook(read_path, write_path, do_not_email_path):
    # Load the 'DO NOT EMAIL' list
    do_not_email_set = load_do_not_email_list(do_not_email_path)
    
    # Initialize ExcelWriter for the new workbook
    writer = pd.ExcelWriter(write_path, engine='openpyxl')

    # Read the Excel file
    df = pd.concat(pd.read_excel(read_path, sheet_name=None), ignore_index=True)

    # Normalize column names
    df.columns = [col.strip().upper().replace(' ', '_') for col in df.columns]

    # Ensure the dataframe contains the required columns
    if 'MERGE_STATUS' in df.columns and 'EMAIL' in df.columns:
        # Remove emails present in the 'DO NOT EMAIL' list
        df = df[~df['EMAIL'].str.lower().isin(do_not_email_set)]

        # Remove duplicate emails
        df.drop_duplicates(subset='EMAIL', inplace=True)

        # Filter out bad emails based on merge status and email validation
        df['VALID_EMAIL'] = df['EMAIL'].apply(is_valid_email)
        bad_email_df = df[(df['MERGE_STATUS'].isin(MERGE_DROP_LIST)) | (~df['VALID_EMAIL'])]
        update_do_not_email_list(bad_email_df['EMAIL'].str.lower(), do_not_email_path)

        # Filter for good emails
        good_email_df = df[(~df['MERGE_STATUS'].isin(MERGE_DROP_LIST)) & (df['VALID_EMAIL'])]

        # Split and save the good emails to new sheets in the new workbook
        chunk_size = 1300
        for i in range(0, len(good_email_df), chunk_size):
            # Inside your loop, modify the good_email_df to select and reorder columns before saving
            good_email_df_filtered = good_email_df[i:i+chunk_size][['COMPANY', 'NAME', 'FIRST', 'EMAIL']]
            good_email_df_filtered.to_excel(writer, sheet_name=str(i//chunk_size + 1), index=False)

    # Save and close the writer
    writer.close()

# Run the function
process_workbook(EXCEL_FILE_PATH, NEW_EXCEL_FILE_PATH, DO_NOT_EMAIL_FILE_PATH)